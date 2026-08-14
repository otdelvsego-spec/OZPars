"""Single user-facing Excel export for analogue comparison reports."""

from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import PriceResult

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def safe_filename(value: str, fallback: str = "OZ_Аналоги") -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "_", value).strip(" ._")
    return cleaned[:120] or fallback


def suggested_export_name(report_name: str) -> str:
    return safe_filename(report_name) + ".xlsx"


def export_results(
    results: list[PriceResult],
    path: str | Path,
    *,
    analogue_limit: int = 5,
) -> Path:
    """Export only the compact analogue table requested by the user."""
    if analogue_limit < 1 or analogue_limit > 30:
        raise ValueError("Количество аналогов для экспорта должно быть от 1 до 30.")

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Аналоги"

    headers = [
        "Исходный артикул",
        "Наименование исходного артикула",
        "Цена исходного артикула, ₽",
    ]
    for position in range(1, analogue_limit + 1):
        headers.extend([f"Аналог {position}", f"Цена аналога {position}, ₽"])
    headers.extend(
        [
            "Медианная цена рынка, ₽",
            "Положение относительно рынка",
            "Отклонение от рынка, %",
        ]
    )
    worksheet.append(headers)

    for result in results:
        row = [
            result.sku,
            result.ozon_name or result.input_name,
            result.comparison_price or result.current_price,
        ]
        for position in range(analogue_limit):
            if position < len(result.analogues):
                analogue = result.analogues[position]
                row.extend(
                    [
                        analogue.name or analogue.sku,
                        analogue.comparison_price or analogue.current_price,
                    ]
                )
            else:
                row.extend([None, None])
        row.extend(
            [
                result.analog_median_price,
                result.market_position,
                result.price_vs_market_percent,
            ]
        )
        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 34
    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    source_price_column = 3
    market_price_column = 4 + analogue_limit * 2
    market_percent_column = market_price_column + 2
    for row_number, result in enumerate(results, start=2):
        worksheet.cell(row_number, source_price_column).number_format = '#,##0.00 "₽"'
        for position in range(analogue_limit):
            name_column = 4 + position * 2
            price_column = name_column + 1
            worksheet.cell(row_number, price_column).number_format = '#,##0.00 "₽"'
            if position < len(result.analogues):
                analogue = result.analogues[position]
                name_cell = worksheet.cell(row_number, name_column)
                if analogue.url:
                    name_cell.hyperlink = analogue.url
                    name_cell.style = "Hyperlink"
        worksheet.cell(row_number, market_price_column).number_format = '#,##0.00 "₽"'
        worksheet.cell(row_number, market_percent_column).number_format = "0.00%"

        source_name_cell = worksheet.cell(row_number, 2)
        if result.url:
            source_name_cell.hyperlink = result.url
            source_name_cell.style = "Hyperlink"

    widths = [18, 38, 23]
    for _ in range(analogue_limit):
        widths.extend([42, 21])
    widths.extend([24, 28, 24])
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(target)
    return target
