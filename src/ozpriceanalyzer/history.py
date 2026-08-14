"""Persistent Ozon price history stored in XLSX."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import PriceResult


class PriceHistoryStore:
    SHEET_NAME = "История"
    HEADERS = (
        "Дата проверки",
        "Артикул Ozon",
        "Название Ozon",
        "Цена, ₽",
        "Цена с Ozon Картой, ₽",
        "Цена без Ozon Карты, ₽",
        "Статус",
        "Ошибка",
    )

    def __init__(self, filename: str | Path) -> None:
        self.path = Path(filename)

    def update(self, results: list[PriceResult], checked_at: datetime | None = None) -> Path:
        timestamp = (checked_at or datetime.now().astimezone()).replace(tzinfo=None, microsecond=0)
        workbook, worksheet = self._open_workbook()
        latest = self._latest_prices(worksheet)
        checked_text = timestamp.strftime("%Y-%m-%d %H:%M:%S")

        for result in results:
            result.checked_at = checked_text
            previous = latest.get(result.sku)
            if result.status == "ok" and result.current_price is not None and previous is not None:
                result.previous_price = previous
                result.price_change = round(result.current_price - previous, 2)
                if previous:
                    result.price_change_percent = result.price_change / previous
            worksheet.append(
                [
                    timestamp,
                    result.sku,
                    result.ozon_name,
                    result.current_price,
                    result.ozon_card_price,
                    result.price_without_card,
                    result.status,
                    result.error,
                ]
            )

        self._format(worksheet)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(self.path)
        return self.path

    def _open_workbook(self):
        if self.path.is_file():
            workbook = load_workbook(self.path)
            if self.SHEET_NAME not in workbook.sheetnames:
                raise ValueError(f"В истории отсутствует лист «{self.SHEET_NAME}».")
            worksheet = workbook[self.SHEET_NAME]
            headers = tuple(worksheet.cell(1, index).value for index in range(1, len(self.HEADERS) + 1))
            if headers != self.HEADERS:
                raise ValueError("Файл истории OZPriceAnalyzer имеет неизвестный формат.")
            return workbook, worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.SHEET_NAME
        worksheet.append(list(self.HEADERS))
        return workbook, worksheet

    @staticmethod
    def _latest_prices(worksheet) -> dict[str, float]:
        latest: dict[str, float] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            sku = "" if row[1] is None else str(row[1]).strip()
            price = row[3]
            status = "" if row[6] is None else str(row[6]).strip()
            if sku and status == "ok" and price is not None:
                try:
                    latest[sku] = float(price)
                except (TypeError, ValueError):
                    pass
        return latest

    @staticmethod
    def _format(worksheet) -> None:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
        widths = (21, 16, 42, 16, 22, 25, 14, 45)
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row, 1).number_format = "dd.mm.yyyy hh:mm:ss"
            for column in (4, 5, 6):
                worksheet.cell(row, column).number_format = '#,##0.00 "₽"'
