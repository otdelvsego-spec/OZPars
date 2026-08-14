"""Excel report generation."""

from __future__ import annotations

from pathlib import Path
from statistics import mean

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .characteristics import format_dimensions, format_materials
from .models import PriceResult


class ExcelReportWriter:
    RESULT_COLUMNS = [
        "Артикул Ozon", "Название из файла", "Название Ozon", "Бренд", "Категория",
        "Габариты, см", "Вес, кг", "Материалы",
        "Цена, ₽", "Цена с Ozon Картой, ₽", "Цена без Ozon Карты, ₽", "Старая цена, ₽",
        "Целевая цена, ₽", "Отклонение от цели, ₽", "Цель достигнута",
        "Предыдущая цена, ₽", "Изменение цены, ₽", "Изменение цены, %",
        "Проверено кандидатов", "Количество аналогов", "Минимальная цена аналогов, ₽",
        "Медианная цена аналогов, ₽", "Средняя цена аналогов, ₽",
        "Аналогов дешевле", "Аналогов дороже", "Отклонение от медианы, ₽",
        "Отклонение от медианы, %", "Позиция относительно рынка",
        "Диагностика аналогов", "Ошибка сравнения", "Ошибка характеристик",
        "Рейтинг", "Отзывы", "Дата проверки", "Ссылка", "Статус", "Ошибка",
    ]

    def __init__(self, filename: str | Path) -> None:
        self.path = Path(filename)

    def write(self, results: list[PriceResult]) -> Path:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        rows = [self._result_row(item) for item in results]
        pd.DataFrame(rows, columns=self.RESULT_COLUMNS).to_excel(
            self.path, sheet_name="Результаты", index=False, engine="openpyxl"
        )
        workbook = load_workbook(self.path)
        self._format_sheet(workbook["Результаты"])
        self._create_summary(workbook, results)
        self._create_analogues(workbook, results)
        workbook.save(self.path)
        return self.path

    @staticmethod
    def _result_row(result: PriceResult) -> list[object]:
        target_reached = None if result.target_reached is None else ("Да" if result.target_reached else "Нет")
        return [
            result.sku, result.input_name, result.ozon_name, result.brand, result.category_name,
            format_dimensions(result.characteristics), result.characteristics.weight_kg,
            format_materials(result.characteristics), result.current_price, result.ozon_card_price,
            result.price_without_card, result.old_price, result.target_price, result.target_difference,
            target_reached, result.previous_price, result.price_change, result.price_change_percent,
            result.analog_checked, result.analog_count, result.analog_min_price,
            result.analog_median_price, result.analog_average_price, result.cheaper_analogs,
            result.more_expensive_analogs, result.price_vs_market, result.price_vs_market_percent,
            result.market_position, result.analog_diagnostics, result.market_error,
            result.characteristics_error, result.rating, result.feedbacks, result.checked_at,
            result.url, result.status, result.error,
        ]

    def _create_summary(self, workbook, results: list[PriceResult]) -> None:
        worksheet = workbook.create_sheet("Сводка")
        successful = [item for item in results if item.status == "ok"]
        compared = [item for item in successful if item.analog_count > 0]
        current_prices = [item.current_price for item in successful if item.current_price is not None]
        metrics = [
            ("Всего артикулов", len(results)),
            ("Успешно обработано", len(successful)),
            ("Сравнение с рынком выполнено", len(compared)),
            ("Ниже рынка", sum(item.market_position == "Ниже рынка" for item in compared)),
            ("На уровне рынка", sum(item.market_position == "На уровне рынка" for item in compared)),
            ("Выше рынка", sum(item.market_position == "Выше рынка" for item in compared)),
            ("Средняя текущая цена, ₽", round(mean(current_prices), 2) if current_prices else None),
            ("Цель достигнута", sum(item.target_reached is True for item in successful)),
            ("Цена снизилась", sum((item.price_change or 0) < 0 for item in successful)),
            ("Цена выросла", sum((item.price_change or 0) > 0 for item in successful)),
        ]
        worksheet.append(["Показатель", "Значение"])
        for metric in metrics:
            worksheet.append(metric)
        self._format_sheet(worksheet)
        worksheet.column_dimensions["A"].width = 38
        worksheet.column_dimensions["B"].width = 22

    def _create_analogues(self, workbook, results: list[PriceResult]) -> None:
        worksheet = workbook.create_sheet("Аналоги")
        headers = [
            "Исходный артикул", "Артикул аналога", "Название", "Бренд", "Категория",
            "Сходство, %", "Сравнено по", "Цена аналога, ₽", "Цена с Ozon Картой, ₽",
            "Цена без Ozon Карты, ₽", "Габариты аналога, см", "Вес аналога, кг",
            "Материалы аналога", "Рейтинг", "Отзывы", "Ссылка",
        ]
        worksheet.append(headers)
        for result in results:
            for analogue in result.analogues:
                worksheet.append(
                    [
                        result.sku, analogue.sku, analogue.name, analogue.brand, analogue.category_name,
                        analogue.similarity_score, ", ".join(analogue.matched_fields),
                        analogue.current_price, analogue.ozon_card_price, analogue.price_without_card,
                        format_dimensions(analogue.characteristics), analogue.characteristics.weight_kg,
                        format_materials(analogue.characteristics), analogue.rating, analogue.feedbacks,
                        analogue.url,
                    ]
                )
        self._format_sheet(worksheet)
        if worksheet.max_row > 1:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row, 6).number_format = "0.00%"
                for column in (8, 9, 10):
                    worksheet.cell(row, column).number_format = '#,##0.00 "₽"'
                cell = worksheet.cell(row, 16)
                if cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"

    @staticmethod
    def _format_sheet(worksheet) -> None:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        fill = PatternFill("solid", fgColor="005BFF")
        font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[1].height = 34
        for column in worksheet.columns:
            letter = column[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column[:50])
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 46)
        for row in range(2, worksheet.max_row + 1):
            for cell in worksheet[row]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        headers = {cell.value: cell.column for cell in worksheet[1]}
        for name in (
            "Цена, ₽", "Цена с Ozon Картой, ₽", "Цена без Ozon Карты, ₽", "Старая цена, ₽",
            "Целевая цена, ₽", "Отклонение от цели, ₽", "Предыдущая цена, ₽",
            "Изменение цены, ₽", "Минимальная цена аналогов, ₽", "Медианная цена аналогов, ₽",
            "Средняя цена аналогов, ₽", "Отклонение от медианы, ₽",
        ):
            if name in headers:
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row, headers[name]).number_format = '#,##0.00 "₽"'
        for name in ("Изменение цены, %", "Отклонение от медианы, %"):
            if name in headers:
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row, headers[name]).number_format = "0.00%"
        if "Ссылка" in headers:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, headers["Ссылка"])
                if cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"
