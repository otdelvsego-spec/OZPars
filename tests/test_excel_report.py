import pandas as pd
from openpyxl import load_workbook

from ozpriceanalyzer.excel import ExcelLoader
from ozpriceanalyzer.models import MarketProduct, PriceResult, ProductCharacteristics
from ozpriceanalyzer.report import ExcelReportWriter


def test_excel_loader_accepts_article_alias(tmp_path) -> None:
    source = tmp_path / "products.xlsx"
    pd.DataFrame([{"Артикул": 3362949093.0, "Название": "Стол", "Целевая цена": "12990"}]).to_excel(source, index=False)
    products = ExcelLoader(source).load()
    assert products[0].sku == "3362949093"
    assert products[0].target_price == 12990


def test_report_creates_three_sheets(tmp_path) -> None:
    characteristics = ProductCharacteristics(
        dimensions_cm={"ширина": 100, "высота": 50},
        weight_kg=10,
        materials=("лдсп", "пластик"),
    )
    analogue = MarketProduct(
        sku="200",
        name="Аналог",
        brand="Brand",
        category_name="Столы",
        current_price=900,
        characteristics=characteristics,
        similarity_score=0.9,
        matched_fields=("материалы 100%", "габариты", "вес"),
        url="https://www.ozon.ru/product/200/",
    )
    result = PriceResult(
        sku="100",
        ozon_name="Исходный",
        current_price=1000,
        characteristics=characteristics,
        analogues=[analogue],
        analog_count=1,
        analog_min_price=900,
        analog_median_price=900,
        analog_average_price=900,
        url="https://www.ozon.ru/product/100/",
    )
    target = tmp_path / "result.xlsx"
    ExcelReportWriter(target).write([result])
    workbook = load_workbook(target)
    assert workbook.sheetnames == ["Результаты", "Сводка", "Аналоги"]
    assert workbook["Аналоги"].max_row == 2
