from openpyxl import load_workbook

from ozpriceanalyzer.exporter import export_results
from ozpriceanalyzer.models import MarketProduct, PriceResult


def test_export_results_creates_only_compact_analogue_sheet(tmp_path) -> None:
    analogue = MarketProduct(
        sku="200",
        name="Аналог",
        brand="",
        category_name="",
        current_price=900.0,
        comparison_price=920.0,
        url="https://www.ozon.ru/product/200/",
    )
    result = PriceResult(
        sku="100",
        ozon_name="Исходный товар",
        current_price=1000.0,
        comparison_price=1050.0,
        analogues=[analogue],
        analog_median_price=920.0,
        market_position="Выше рынка",
        price_vs_market_percent=0.1413,
        url="https://www.ozon.ru/product/100/",
    )

    path = export_results([result], tmp_path / "report.xlsx", analogue_limit=2)
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Аналоги"]
    sheet = workbook["Аналоги"]
    headers = [cell.value for cell in sheet[1]]
    assert headers == [
        "Исходный артикул",
        "Наименование исходного артикула",
        "Цена исходного артикула, ₽",
        "Аналог 1",
        "Цена аналога 1, ₽",
        "Аналог 2",
        "Цена аналога 2, ₽",
        "Медианная цена рынка, ₽",
        "Положение относительно рынка",
        "Отклонение от рынка, %",
    ]
    assert sheet.cell(2, 2).hyperlink.target.endswith("/100/")
    assert sheet.cell(2, 4).hyperlink.target.endswith("/200/")
    assert sheet.cell(2, 5).value == 920.0
    assert sheet.cell(2, 8).value == 920.0
    assert sheet.cell(2, 9).value == "Выше рынка"
